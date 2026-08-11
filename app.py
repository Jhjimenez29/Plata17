from io import BytesIO
import base64
import datetime
import os
import pandas as pd
import requests
import streamlit as st
import unicodedata

# Librerías para formato avanzado de Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# 1. CONFIGURACIÓN E INICIALIZACIÓN
# ==========================================
st.set_page_config(page_title="Catálogo & Auditoría", page_icon="📱", layout="wide")

CLAVE_SUPERVISOR = "super123"

# Estilos CSS
st.markdown("""
    <style>
    .sombra-tenue {
        background-color: #F8FAFC;
        border: 1px solid #E2E8F0;
        padding: 10px 14px;
        border-radius: 6px;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.05);
        margin-bottom: 12px;
    }
    .titulo-negro {
        color: #000000 !important;
        font-weight: 600;
        margin: 0;
    }
    .alerta-ultima-auditoria {
        background-color: #FEF3C7;
        border: 1px solid #F59E0B;
        color: #92400E;
        padding: 8px 12px;
        border-radius: 6px;
        font-size: 13px;
        margin-top: 8px;
    }
    
    /* Personalización de Radio Buttons a color ROJO */
    div[data-testid="stRadio"] label[data-baseweb="radio"] div:first-child {
        background-color: white !important;
    }
    div[data-testid="stRadio"] input[type="radio"]:checked + div {
        border-color: #DC2626 !important;
        background-color: #DC2626 !important;
    }
    div[data-testid="stRadio"] input[type="radio"]:checked + div > div {
        background-color: white !important;
    }
    div[data-testid="stRadio"] input[type="radio"] {
        accent-color: #DC2626 !important;
    }
    </style>
""", unsafe_allow_html=True)

def inicializar_estado_sesion():
    """Inicializa todas las variables globales de sesión"""
    defaults = {
        "pantalla": "login",
        "vista_catalogo": None,
        "filtro_familia": "-- Selecciona una familia --",
        "busqueda_rapida": "",
        "tipo_cliente_seleccion": "Uso libre / Consulta",
        "cliente_nombre": "",
        "cliente_numero": "",
        "productos_seleccionados": {},
        "marcas_seleccionadas": {},
        "target_eliminar_historial": []  # Guarda los índices que se van a eliminar
    }
    for key, val in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = val

inicializar_estado_sesion()

# ==========================================
# 2. FUNCIONES DE UTILIDAD Y LIMPIEZA TOTAL
# ==========================================
def normalizar_texto(texto):
    if not isinstance(texto, str):
        return ""
    texto = texto.strip().lower()
    return ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')

def reiniciar_pantalla_total():
    """Limpia la selección de productos, marcas, cliente y restablece los controles visuales"""
    keys_a_borrar = [
        k for k in st.session_state.keys() 
        if k.startswith("chk_p_") or k.startswith("chk_m_") or 
           k.startswith("sel_ubic_") or k.startswith("sel_ubi_m_") or 
           k in ["radio_modo_consulta", "radio_tipo_cliente"]
    ]
    for key in keys_a_borrar:
        del st.session_state[key]
    
    st.session_state.productos_seleccionados = {}
    st.session_state.marcas_seleccionadas = {}
    st.session_state.tipo_cliente_seleccion = "Uso libre / Consulta"
    st.session_state.cliente_nombre = ""
    st.session_state.cliente_numero = ""
    st.session_state.vista_catalogo = None
    st.session_state.filtro_familia = "-- Selecciona una familia --"
    st.session_state.busqueda_rapida = ""
    st.session_state.target_eliminar_historial = []

# ==========================================
# 3. GESTIÓN DE ARCHIVOS Y PERSISTENCIA (CSV/DATOS)
# ==========================================
@st.cache_data
def cargar_productos():
    for enc in ["utf-8", "latin1"]:
        try:
            return pd.read_csv("productos.csv", encoding=enc)
        except Exception:
            continue
    return None

@st.cache_data
def cargar_lista_precios():
    for enc in ["utf-8", "latin1"]:
        try:
            df = pd.read_csv("lista_precios.csv", encoding=enc)
            if df is not None and not df.empty:
                return df
        except Exception:
            continue
    return None

@st.cache_data
def cargar_marcas():
    for n in ["marcas.csv", "Marcas.csv", "marcas.xlsx", "Marcas.xlsx"]:
        if n.endswith(".csv"):
            for enc in ["utf-8", "latin1"]:
                try:
                    df = pd.read_csv(n, encoding=enc)
                    if df is not None and not df.empty:
                        return df
                except Exception:
                    continue
        elif n.endswith(".xlsx"):
            try:
                df = pd.read_excel(n)
                if df is not None and not df.empty:
                    return df
            except Exception:
                continue
    return None

def cargar_catalogo_clientes():
    archivo = "clientes_directorio.csv"
    if os.path.exists(archivo):
        try:
            df = pd.read_csv(archivo, encoding='utf-8', dtype=str)
            df["Numero Cliente"] = df["Numero Cliente"].fillna("").astype(str).str.strip()
            df["Nombre Cliente"] = df["Nombre Cliente"].fillna("").astype(str).str.strip()
            return df
        except Exception:
            pass
    return pd.DataFrame(columns=["Numero Cliente", "Nombre Cliente"])

# --- RESPALDO AUTOMÁTICO A GITHUB ---
def respaldar_archivo_en_github(nombre_archivo, mensaje_commit):
    """Sube el archivo local (ya guardado en disco) al repositorio de GitHub como respaldo.
    Si no hay credenciales configuradas en Secrets, no hace nada (no rompe la app)."""
    try:
        token = st.secrets["GITHUB_TOKEN"]
        repo = st.secrets["GITHUB_REPO"]
    except Exception:
        return

    try:
        with open(nombre_archivo, "rb") as f:
            contenido_b64 = base64.b64encode(f.read()).decode("utf-8")

        url = f"https://api.github.com/repos/{repo}/contents/{nombre_archivo}"
        headers = {
            "Authorization": f"Bearer {token}",
            "Accept": "application/vnd.github+json"
        }

        # Obtenemos el SHA actual del archivo en GitHub (requerido para poder actualizarlo)
        sha_actual = None
        resp_get = requests.get(url, headers=headers, timeout=10)
        if resp_get.status_code == 200:
            sha_actual = resp_get.json().get("sha")

        payload = {"message": mensaje_commit, "content": contenido_b64}
        if sha_actual:
            payload["sha"] = sha_actual

        requests.put(url, headers=headers, json=payload, timeout=10)
    except Exception:
        # Si el respaldo falla (ej. sin internet), la app sigue funcionando con el archivo local
        pass


def guardar_catalogo_clientes_completo(df):
    df.to_csv("clientes_directorio.csv", index=False, encoding='utf-8')
    respaldar_archivo_en_github("clientes_directorio.csv", "Respaldo automatico: clientes_directorio.csv")

def guardar_cliente_directorio(num_cliente, nombre_cliente):
    num, nom = str(num_cliente).strip(), str(nombre_cliente).strip()
    if not num or not nom:
        return
        
    df_existente = cargar_catalogo_clientes()
    if not df_existente.empty:
        coincidencias = df_existente[
            (df_existente["Numero Cliente"].str.lower() == num.lower()) & 
            (df_existente["Nombre Cliente"].str.lower() == nom.lower())
        ]
        if not coincidencias.empty:
            return
            
    nuevo_df = pd.DataFrame([{"Numero Cliente": num, "Nombre Cliente": nom}])
    df_final = pd.concat([df_existente, nuevo_df], ignore_index=True)
    guardar_catalogo_clientes_completo(df_final)

def cargar_historial_maestro():
    archivo = "historial_revisiones.csv"
    if os.path.exists(archivo):
        try:
            df = pd.read_csv(archivo, encoding='utf-8', dtype=str)
            return df
        except Exception:
            pass
    return pd.DataFrame()

def guardar_historial_maestro_completo(df):
    archivo = "historial_revisiones.csv"
    df.to_csv(archivo, index=False, encoding='utf-8')
    respaldar_archivo_en_github(archivo, "Respaldo automatico: historial_revisiones.csv")

def guardar_en_historial_maestro(df_nuevas_filas):
    archivo = "historial_revisiones.csv"
    header = not os.path.exists(archivo)
    df_nuevas_filas.to_csv(archivo, mode='a', header=header, index=False, encoding='utf-8')
    respaldar_archivo_en_github(archivo, "Respaldo automatico: nueva auditoria en historial_revisiones.csv")

def obtener_ultima_auditoria(nombre_cliente, num_cliente):
    df_h = cargar_historial_maestro()
    if df_h.empty:
        return None
    condicion = pd.Series(False, index=df_h.index)
    if nombre_cliente and "Nombre Cliente" in df_h.columns:
        condicion |= (df_h["Nombre Cliente"].astype(str).str.strip().str.lower() == str(nombre_cliente).strip().lower())
    if num_cliente and "Numero Cliente" in df_h.columns:
        condicion |= (df_h["Numero Cliente"].astype(str).str.strip() == str(num_cliente).strip())
        
    df_cliente = df_h[condicion]
    if not df_cliente.empty:
        return df_cliente["Fecha_Hora"].iloc[-1] if "Fecha_Hora" in df_cliente.columns else df_cliente["Fecha"].iloc[-1]
    return None

def consolidar_visita_actual():
    nombre_c = st.session_state.cliente_nombre or "Cliente General"
    num_c = st.session_state.cliente_numero or "1001"
    
    filas = []
    fecha_std = datetime.date.today().strftime("%Y-%m-%d")
    fecha_hora_actual = datetime.datetime.now().strftime("%d de %B a las %I:%M %p").lower()
    
    for k, v in st.session_state.productos_seleccionados.items():
        filas.append({
            "Fecha": fecha_std,
            "Fecha_Hora": fecha_hora_actual,
            "Numero Cliente": num_c,
            "Nombre Cliente": nombre_c,
            "Esquema": "Esquema Comercial 2017",
            "Tipo Registro": "PRODUCTO",
            "Identificador / Código": k,
            "Descripción / Detalle": v["datos"].get("Descripcion de producto", str(v["datos"])),
            "Ubicación": v["ubicacion"]
        })

    for k, v in st.session_state.marcas_seleccionadas.items():
        filas.append({
            "Fecha": fecha_std,
            "Fecha_Hora": fecha_hora_actual,
            "Numero Cliente": num_c,
            "Nombre Cliente": nombre_c,
            "Esquema": "Esquema Comercial 2017",
            "Tipo Registro": "MARCA",
            "Identificador / Código": k,
            "Descripción / Detalle": f"Exhibidor / Marca: {k}",
            "Ubicación": v["ubicacion"]
        })

    return pd.DataFrame(filas)

def consolidar_y_guardar_visita_actual():
    df_rep = consolidar_visita_actual()
    if not df_rep.empty:
        guardar_en_historial_maestro(df_rep)
        if st.session_state.cliente_nombre and st.session_state.cliente_numero:
            guardar_cliente_directorio(st.session_state.cliente_numero, st.session_state.cliente_nombre)
        return True
    return False

# ==========================================
# 4. MOTOR DE GENERACIÓN DE EXCEL PROFESIONAL
# ==========================================
def generar_excel_profesional(df, titulo_reporte="REPORTE DE AUDITORÍA DE CAMPO"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte_Auditoria"
    ws.views.sheetView[0].showGridLines = True

    fuente_titulo = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    fuente_subtitulo = Font(name="Calibri", size=11, italic=True, color="D9D9D9")
    fuente_meta_bold = Font(name="Calibri", size=10, bold=True, color="1F497D")
    fuente_meta_val = Font(name="Calibri", size=10, color="000000")
    fuente_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fuente_datos = Font(name="Calibri", size=10, color="000000")

    fill_header_principal = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_header_tabla = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_cebra = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")

    borde_fino = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    ws.merge_cells('A1:F1')
    ws['A1'] = f"  {titulo_reporte.upper()}"
    ws['A1'].font = fuente_titulo
    ws['A1'].fill = fill_header_principal
    ws['A1'].alignment = Alignment(vertical='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = f"  Generado el: {datetime.datetime.now().strftime('%d/%m/%Y a las %H:%M hrs')}"
    ws['A2'].font = fuente_subtitulo
    ws['A2'].fill = fill_header_principal
    ws['A2'].alignment = Alignment(vertical='center')

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18

    if "Nombre Cliente" in df.columns and not df["Nombre Cliente"].empty:
        nombre_cli = df["Nombre Cliente"].iloc[0]
    else:
        nombre_cli = st.session_state.cliente_nombre or "Cliente General"

    if "Numero Cliente" in df.columns and not df["Numero Cliente"].empty:
        num_cli = df["Numero Cliente"].iloc[0]
    else:
        num_cli = st.session_state.cliente_numero or "N/A"

    total_prod = len(df[df["Tipo Registro"] == "PRODUCTO"]) if "Tipo Registro" in df.columns else 0
    total_marcas = len(df[df["Tipo Registro"] == "MARCA"]) if "Tipo Registro" in df.columns else 0

    metadatos = [
        ("Cliente:", nombre_cli, "Número Cliente:", num_cli),
        ("Total Productos:", total_prod, "Total Marcas:", total_marcas)
    ]

    row_idx = 4
    for r in metadatos:
        ws.cell(row=row_idx, column=1, value=r[0]).font = fuente_meta_bold
        ws.cell(row=row_idx, column=2, value=r[1]).font = fuente_meta_val
        ws.cell(row=row_idx, column=4, value=r[2]).font = fuente_meta_bold
        ws.cell(row=row_idx, column=5, value=r[3]).font = fuente_meta_val
        row_idx += 1

    start_row = 7
    df_export = df.drop(columns=["Fecha_Hora"], errors="ignore")
    
    for col_idx, col_name in enumerate(df_export.columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=str(col_name))
        cell.font = fuente_header
        cell.fill = fill_header_tabla
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[start_row].height = 22

    for r_idx, row_data in enumerate(df_export.itertuples(index=False), start=start_row + 1):
        ws.row_dimensions[r_idx].height = 20
        usar_cebra = (r_idx % 2 == 0)
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = fuente_datos
            cell.border = borde_fino
            cell.alignment = Alignment(vertical='center')
            if usar_cebra:
                cell.fill = fill_cebra

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row >= start_row and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

# ==========================================
# 5. CARGA Y DETECCIÓN AUTOMÁTICA DE COLUMNAS
# ==========================================
df_productos = cargar_productos()
df_marcas = cargar_marcas()

columna_familia_real = None
columna_codigo_real = None
columna_clave_real = None
columna_num_fam_real = None
columna_desc_prod_real = None

if df_productos is not None:
    df_productos.columns = df_productos.columns.str.strip()
    for col in df_productos.columns:
        col_norm = normalizar_texto(col)
        if "familia" in col_norm and "numero" not in col_norm and "num" not in col_norm:
            columna_familia_real = col
        elif col_norm in ["codigo", "cod", "sku", "id"]:
            columna_codigo_real = col
        elif col_norm in ["clave", "clv"]:
            columna_clave_real = col
        elif "numero de familia" in col_norm or "num de familia" in col_norm or "num familia" in col_norm or "n_familia" in col_norm:
            columna_num_fam_real = col
        elif "descripcion" in col_norm or "desc" in col_norm or "producto" in col_norm:
            if not columna_desc_prod_real:
                columna_desc_prod_real = col

    if not columna_familia_real:
        fam_cols = [c for c in df_productos.columns if "familia" in normalizar_texto(c)]
        if fam_cols:
            columna_familia_real = fam_cols[0]

    if not columna_desc_prod_real and len(df_productos.columns) >= 2:
        columna_desc_prod_real = df_productos.columns[1]

    renombrar_dict = {}
    if columna_num_fam_real:
        renombrar_dict[columna_num_fam_real] = "Numero de familia"
    if columna_desc_prod_real:
        renombrar_dict[columna_desc_prod_real] = "Descripcion de producto"
        
    if renombrar_dict:
        df_productos = df_productos.rename(columns=renombrar_dict)

    st.session_state.columnas_seleccionadas = [c for c in df_productos.columns if c not in ["Fecha", "Fecha_Hora", "Código", "Clave", "DescFam"]]

# ==========================================
# 6. VISTAS DE LA APLICACIÓN
# ==========================================

# --- PANTALLA: LOGIN ---
if st.session_state.pantalla == "login":
    st.markdown("<h1 style='text-align: center;'>🖼️</h1>", unsafe_allow_html=True)
    st.markdown("<h2 style='text-align: center; color: #000000;'>Visualizador de Catálogo & Auditoría</h2>", unsafe_allow_html=True)
    st.markdown("---")

    col_l1, col_l2, col_l3 = st.columns([1, 2, 1])
    with col_l2:
        usuario = st.text_input("Usuario", placeholder="Introduce tu usuario", key="input_user")
        contrasena = st.text_input("Contraseña", type="password", placeholder="Introduce tu contraseña", key="input_pass")
        st.write("")
        if st.button("Ingresar", use_container_width=True, type="primary"):
            if usuario == "admin" and contrasena == "1234":
                st.session_state.pantalla = "menu_principal"
                st.rerun()
            else:
                st.error("Usuario o contraseña incorrectos.")

# --- PANTALLA: MENÚ PRINCIPAL ---
elif st.session_state.pantalla == "menu_principal":
    st.markdown("<h2 style='text-align: center; color: #000000;'>Auditorías</h2>", unsafe_allow_html=True)
    st.markdown("---")

    col_m1, col_m2, col_m3 = st.columns([1, 4, 1])
    with col_m2:
        if st.button("📋 Auditorías", use_container_width=True, type="primary"):
            st.session_state.pantalla = "resultados"
            st.rerun()
        st.write("")
        if st.button("💲 Hd", use_container_width=True, type="primary"):
            st.session_state.pantalla = "hd_consulta"
            st.rerun()
        st.write("")
        if st.button("🚪 Cerrar Sesión", use_container_width=True):
            st.session_state.pantalla = "login"
            st.rerun()

# --- PANTALLA: HD (CONSULTA DE PRECIOS) ---
elif st.session_state.pantalla == "hd_consulta":
    col_hd1, col_hd2 = st.columns([7, 3])
    with col_hd1:
        st.markdown("<h3 style='margin:0;'>💲 Hd</h3>", unsafe_allow_html=True)
    with col_hd2:
        if st.button("⬅ Regresar al Menú Principal", use_container_width=True):
            st.session_state.pantalla = "menu_principal"
            st.rerun()

    st.markdown("---")

    df_precios_hd = cargar_lista_precios()

    if df_precios_hd is None:
        st.warning("⚠️ No se cargó 'lista_precios.csv'.")
    elif "HD" not in df_precios_hd.columns:
        st.warning("⚠️ El archivo 'lista_precios.csv' todavía no tiene la columna 'HD'.")
    else:
        df_hd = df_precios_hd[df_precios_hd["HD"].notna() & (df_precios_hd["HD"].astype(str).str.strip() != "")].copy()

        if "precio mayoreo con IVA" in df_hd.columns:
            df_hd["precio mayoreo con IVA"] = df_hd["precio mayoreo con IVA"].apply(
                lambda x: f"$ {float(x):,.2f}" if pd.notna(x) and str(x).strip() != "" else ""
            )

        columnas_hd_visibles = [c for c in ["código", "precio mayoreo con IVA", "descripción", "clave"] if c in df_hd.columns]

        modo_hd = st.radio("Modo:", options=["🔍 Búsqueda", "📋 Ver Lista Completa"], label_visibility="collapsed", horizontal=True)

        if modo_hd == "🔍 Búsqueda":
            columnas_busqueda_hd = [c for c in ["código", "clave", "descripción"] if c in df_hd.columns]
            texto_busqueda_hd = st.text_input("🔍 Búsqueda:", placeholder="Escribe código, clave o descripción...", key="input_busqueda_hd")

            if texto_busqueda_hd.strip():
                mascara_hd = pd.Series(False, index=df_hd.index)
                for col in columnas_busqueda_hd:
                    mascara_hd = mascara_hd | df_hd[col].astype(str).str.contains(texto_busqueda_hd.strip(), case=False, na=False)
                df_resultado_hd = df_hd[mascara_hd].sort_values(by="descripción", ascending=True) if "descripción" in df_hd.columns else df_hd[mascara_hd]

                st.caption(f"Coincidencias: `{len(df_resultado_hd)}`")
                st.dataframe(df_resultado_hd[columnas_hd_visibles], use_container_width=True, hide_index=True, height=400)
            else:
                st.info("Escribe algo arriba para buscar entre los productos Hd.")
        else:
            df_lista_hd = df_hd.sort_values(by="descripción", ascending=True) if "descripción" in df_hd.columns else df_hd
            st.caption(f"Total de productos Hd: `{len(df_lista_hd)}`")
            st.dataframe(df_lista_hd[columnas_hd_visibles], use_container_width=True, hide_index=True, height=500)

# --- PANTALLA: GESTIÓN DE CLIENTES ---
elif st.session_state.pantalla == "gestion_clientes":
    col_g1, col_g2 = st.columns([7, 3])
    with col_g1:
        st.markdown("<h3 style='margin:0;'>👥 Directorio, Alta y Modificación de Clientes</h3>", unsafe_allow_html=True)
    with col_g2:
        if st.button("← Volver a Búsqueda", use_container_width=True, type="primary"):
            st.session_state.pantalla = "resultados"
            st.rerun()

    st.markdown("---")
    col_alta1, col_alta2 = st.columns([4, 6])

    with col_alta1:
        st.markdown("#### ➕ Registrar Nuevo Cliente")
        with st.form("form_alta_cliente", clear_on_submit=True):
            n_cli = st.text_input("Número / Código de Cliente:")
            nom_cli = st.text_input("Nombre / Razón Social:")
            if st.form_submit_button("💾 Guardar Registro", use_container_width=True, type="primary"):
                if n_cli and nom_cli:
                    guardar_cliente_directorio(n_cli, nom_cli)
                    st.success(f"✅ ¡Registro '{nom_cli}' ({n_cli}) agregado con éxito!")
                    st.rerun()
                else:
                    st.error("⚠️ Debes completar número y nombre del cliente.")

    with col_alta2:
        st.markdown("#### 📁 Directorio de Clientes")
        df_dir = cargar_catalogo_clientes()
        if df_dir.empty:
            st.info("ℹ️ Aún no hay clientes registrados.")
        else:
            tab_ver, tab_editar, tab_eliminar = st.tabs(["👁️ Ver Directorio", "✏️ Editar Registro", "🗑️ Eliminar"])
            
            with tab_ver:
                st.dataframe(df_dir, use_container_width=True, hide_index=True, height=280)

            with tab_editar:
                opciones = [f"{i} - {r['Nombre Cliente']} {r['Numero Cliente']}" for i, r in df_dir.iterrows()]
                sel = st.selectbox("Selecciona cliente a modificar:", options=["-- Seleccionar --"] + opciones)
                if sel != "-- Seleccionar --":
                    idx = int(sel.split(" - ")[0])
                    r = df_dir.iloc[idx]
                    with st.form("form_edit"):
                        num_m = st.text_input("Nuevo Número:", value=str(r["Numero Cliente"]))
                        nom_m = st.text_input("Nuevo Nombre:", value=str(r["Nombre Cliente"]))
                        if st.form_submit_button("💾 Guardar Cambios", use_container_width=True, type="primary"):
                            if num_m and nom_m:
                                df_dir.at[idx, "Numero Cliente"] = str(num_m).strip()
                                df_dir.at[idx, "Nombre Cliente"] = str(nom_m).strip()
                                guardar_catalogo_clientes_completo(df_dir)
                                st.success("✅ Registro actualizado!")
                                st.rerun()

            with tab_eliminar:
                opciones_del = [f"{i} - {r['Nombre Cliente']} {r['Numero Cliente']}" for i, r in df_dir.iterrows()]
                sel_del = st.selectbox("Selecciona cliente a eliminar:", options=["-- Seleccionar --"] + opciones_del)
                if sel_del != "-- Seleccionar --":
                    idx_d = int(sel_del.split(" - ")[0])
                    st.warning(f"⚠️ Estás por eliminar: **{df_dir.iloc[idx_d]['Nombre Cliente']}**")
                    clave = st.text_input("🔒 Clave de Supervisor:", type="password")
                    if st.button("🚨 Confirmar Eliminación", type="primary", use_container_width=True):
                        if clave == CLAVE_SUPERVISOR:
                            df_dir = df_dir.drop(index=idx_d).reset_index(drop=True)
                            guardar_catalogo_clientes_completo(df_dir)
                            st.success("✅ Registro eliminado.")
                            st.rerun()
                        else:
                            st.error("❌ Clave incorrecta.")

# --- PANTALLA: REPORTE DE LA VISITA ACTUAL ---
elif st.session_state.pantalla == "reporte_auditoria":
    col_nav1, col_nav2 = st.columns([7, 3])
    with col_nav1:
        st.markdown("<h3 style='color: #000000; margin:0;'>📋 Reporte de Visita Actual</h3>", unsafe_allow_html=True)
    with col_nav2:
        if st.button("← Volver a Búsqueda", use_container_width=True, type="primary"):
            st.session_state.pantalla = "resultados"
            st.rerun()

    st.markdown("---")
    df_reporte_actual = consolidar_visita_actual()

    if df_reporte_actual.empty:
        st.info("ℹ️ Aún no has seleccionado productos ni marcas para este cliente.")
    else:
        st.markdown("#### 👤 Datos del Cliente / Revisión")
        col_c1, col_c2, col_c3, col_c4 = st.columns(4)
        with col_c1:
            st.session_state.cliente_nombre = st.text_input("Nombre del Cliente:", value=st.session_state.cliente_nombre or "Cliente General")
        with col_c2:
            st.session_state.cliente_numero = st.text_input("Número de Cliente:", value=st.session_state.cliente_numero or "1001")
        with col_c3:
            st.date_input("Fecha de Revisión:", datetime.date.today())
        with col_c4:
            st.text_input("Esquema:", value="Esquema Comercial 2017", disabled=True)

        st.markdown("---")
        st.dataframe(df_reporte_actual.drop(columns=["Fecha_Hora"], errors="ignore"), use_container_width=True, hide_index=True, height=300)

        st.markdown("---")
        col_acc1, col_acc2, col_acc3 = st.columns(3)
        
        with col_acc1:
            if st.button("💾 Confirmar y Guardar Visita Actual", use_container_width=True, type="primary"):
                nombre_guardado = st.session_state.cliente_nombre or "Cliente General"
                if consolidar_y_guardar_visita_actual():
                    st.toast(f"✅ ¡Visita de '{nombre_guardado}' guardada con éxito!")
                reiniciar_pantalla_total()
                st.session_state.pantalla = "resultados"
                st.rerun()

        with col_acc2:
            bytes_excel = generar_excel_profesional(df_reporte_actual, titulo_reporte="REPORTE DE VISITA Y AUDITORÍA")
            st.download_button(
                label="📥 Descargar Reporte en Excel Pro",
                data=bytes_excel,
                file_name=f"Visita_{st.session_state.cliente_numero}_{datetime.date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        with col_acc3:
            if st.button("🗑️ Descartar Selección Actual", use_container_width=True):
                reiniciar_pantalla_total()
                st.session_state.pantalla = "resultados"
                st.rerun()

# --- PANTALLA: HISTORIAL GENERAL Y REPORTES ---
elif st.session_state.pantalla == "historial":
    col_h1, col_h2 = st.columns([7, 3])
    with col_h1:
        st.markdown("<h3 style='margin:0;'>📊 Historial de Revisiones Realizadas</h3>", unsafe_allow_html=True)
    with col_h2:
        if st.button("← Volver a Búsqueda", use_container_width=True, type="primary"):
            st.session_state.pantalla = "resultados"
            st.rerun()

    st.markdown("---")
    df_historial = cargar_historial_maestro()

    if df_historial.empty:
        st.info("ℹ️ Aún no hay visitas guardadas en el historial.")
    else:
        st.markdown("#### 🔍 Filtros para Consulta y Edición")
        col_f1, col_f2, col_f3 = st.columns(3)
        opcion_default = "-- Seleccionar --"
        opcion_todos = "Todos"
        
        opciones_clientes = [opcion_default, opcion_todos]
        if "Nombre Cliente" in df_historial.columns and "Numero Cliente" in df_historial.columns:
            clientes_unicos = df_historial[["Nombre Cliente", "Numero Cliente"]].drop_duplicates()
            for _, r in clientes_unicos.iterrows():
                nom = str(r["Nombre Cliente"]).strip()
                num = str(r["Numero Cliente"]).strip()
                if nom or num:
                    opciones_clientes.append(f"{num} - {nom}")
        
        with col_f1:
            cliente_sel_str = st.selectbox("Filtrar por Cliente (Número o Nombre):", options=opciones_clientes)
        
        with col_f2:
            fechas_opciones = sorted(df_historial["Fecha"].dropna().unique().tolist(), reverse=True) if "Fecha" in df_historial.columns else []
            fecha_sel = st.selectbox("Filtrar por Fecha:", options=[opcion_default] + fechas_opciones)

        with col_f3:
            busqueda_num_cli = st.text_input("O busca directamente por Número de Cliente:", placeholder="Ej: 588816")

        # Aplicar filtros manteniendo los índices originales para el borrado seguro
        if cliente_sel_str == opcion_default and not busqueda_num_cli.strip():
            df_filtrado = pd.DataFrame(columns=df_historial.columns)
        elif cliente_sel_str == opcion_todos:
            df_filtrado = df_historial.copy()
        elif cliente_sel_str != opcion_default:
            num_extraido = cliente_sel_str.split(" - ")[0].strip()
            df_filtrado = df_historial[df_historial["Numero Cliente"].astype(str).str.strip() == num_extraido].copy()
        else:
            df_filtrado = df_historial.copy()

        if busqueda_num_cli.strip():
            df_filtrado = df_historial[df_historial["Numero Cliente"].astype(str).str.contains(busqueda_num_cli.strip(), case=False, na=False)].copy()

        if fecha_sel != opcion_default and not df_filtrado.empty:
            df_filtrado = df_filtrado[df_filtrado["Fecha"] == fecha_sel]

        # Despliegue en pestañas
        if cliente_sel_str == opcion_default and not busqueda_num_cli.strip():
            pass  # Estado limpio inicial
        elif not df_filtrado.empty:
            st.markdown("---")
            
            tab_vista = st.container()

            # --- VISTA ÚNICA CON EDICIÓN/ELIMINACIÓN DIRECTA EN TABLA NATIVA ---
            with tab_vista:
                # Contenedor superior ubicado al INICIO de la tabla
                container_superior = st.container()

                # Preparamos la tabla con las dos columnas de casillas integradas
                df_editor = df_filtrado.drop(columns=["Fecha_Hora"], errors="ignore").copy()
                df_editor["Seleccionar"] = False        # Casilla para borrado múltiple
                df_editor["🗑️ Eliminar"] = False        # Casilla para borrado individual inmediato

                edited_df = st.data_editor(
                    df_editor,
                    column_config={
                        "Seleccionar": st.column_config.CheckboxColumn(
                            "Seleccionar",
                            help="Marca para eliminar junto con otros registros seleccionados",
                            default=False
                        ),
                        "🗑️ Eliminar": st.column_config.CheckboxColumn(
                            "Eliminar",
                            help="Marca para eliminar ÚNICAMENTE este registro de inmediato",
                            default=False
                        )
                    },
                    disabled=[c for c in df_editor.columns if c not in ["Seleccionar", "🗑️ Eliminar"]],
                    hide_index=True,
                    use_container_width=True,
                    height=320,
                    key="editor_tabla_historial_vista"
                )

                # Capturar las filas marcadas para borrado múltiple (columna "Seleccionar")
                filas_seleccionadas_indices = df_filtrado.index[edited_df["Seleccionar"]].tolist()

                # Capturar la fila marcada para borrado individual inmediato (columna "Eliminar")
                filas_accion_individual = df_filtrado.index[edited_df["🗑️ Eliminar"]].tolist()
                if filas_accion_individual:
                    st.session_state.target_eliminar_historial = filas_accion_individual

                # Llenamos el contenedor superior (al inicio de la tabla)
                with container_superior:
                    col_top_v1, col_top_v2 = st.columns([5, 5])
                    
                    # 1. Botón "Descargar reporte en excel" a la IZQUIERDA
                    with col_top_v1:
                        num_cli_f = "TODOS" if cliente_sel_str == opcion_todos else (df_filtrado['Numero Cliente'].iloc[0] if 'Numero Cliente' in df_filtrado.columns else 'General')
                        bytes_excel_h = generar_excel_profesional(df_filtrado, titulo_reporte=f"REPORTE DE HISTORIAL - CLIENTE {num_cli_f}")
                        st.download_button(
                            label="📥 Descargar reporte en Excel",
                            data=bytes_excel_h,
                            file_name=f"Reporte_Cliente_{num_cli_f}_{datetime.date.today()}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            type="primary"
                        )
                        st.caption(f"Registros mostrados: `{len(df_filtrado)}` filas")

                    # 3. Botón "Eliminar registros seleccionados" al inicio pero a la DERECHA
                    with col_top_v2:
                        if len(filas_seleccionadas_indices) > 0:
                            if st.button(f"🚨 Eliminar ({len(filas_seleccionadas_indices)}) Registro(s) Seleccionado(s)", type="primary", use_container_width=True):
                                st.session_state.target_eliminar_historial = filas_seleccionadas_indices
                                st.rerun()

            # --- PANEL DE SEGURIDAD / CONFIRMACIÓN Y CLAVE DE SUPERVISOR (VENTANA EMERGENTE) ---
            @st.dialog("⚠️ Confirmación de Eliminación")
            def dialogo_confirmar_eliminacion_historial():
                num_a_borrar = len(st.session_state.target_eliminar_historial)
                st.warning(f"¿Estás seguro de que deseas eliminar permanentemente **{num_a_borrar}** registro(s) del historial maestro?")

                clave_elim = st.text_input("🔒 Introduce Clave de Supervisor:", type="password", key="pass_del_hist_modal")

                col_p2, col_p3 = st.columns(2)
                with col_p2:
                    if st.button("💥 Confirmar Eliminación", type="primary", use_container_width=True):
                        if clave_elim == CLAVE_SUPERVISOR:
                            df_hist_actualizado = df_historial.drop(index=st.session_state.target_eliminar_historial).reset_index(drop=True)
                            guardar_historial_maestro_completo(df_hist_actualizado)
                            st.session_state.target_eliminar_historial = []
                            st.session_state.pop("editor_tabla_historial_vista", None)
                            st.success(f"✅ ¡Se han eliminado {num_a_borrar} registro(s) correctamente!")
                            st.rerun()
                        else:
                            st.error("❌ Clave de supervisor incorrecta.")
                with col_p3:
                    if st.button("❌ Cancelar", use_container_width=True):
                        st.session_state.target_eliminar_historial = []
                        st.session_state.pop("editor_tabla_historial_vista", None)
                        st.rerun()

            if len(st.session_state.target_eliminar_historial) > 0:
                dialogo_confirmar_eliminacion_historial()
        else:
            st.markdown("---")
            st.warning("⚠️ No se encontraron registros que coincidan con la búsqueda.")

# --- PANTALLA PRINCIPAL: BÚSQUEDA Y AUDITORÍA ---
elif st.session_state.pantalla == "resultados":
    total_sel = len(st.session_state.productos_seleccionados) + len(st.session_state.marcas_seleccionadas)
    
    col_sup1, col_sup2 = st.columns([7, 3])
    with col_sup2:
        col_b3, col_b4 = st.columns([1.5, 2])
        with col_b3:
            with st.popover("⋮", use_container_width=True):
                lbl = f"📋 Ver Visita ({total_sel})" if total_sel > 0 else "📋 Ver Visita"
                if st.button(lbl, use_container_width=True, type="primary" if total_sel > 0 else "secondary"):
                    st.session_state.pantalla = "reporte_auditoria"
                    st.rerun()
                if st.button("📊 Historial", use_container_width=True):
                    st.session_state.pantalla = "historial"
                    st.rerun()
                if st.button("👤 Gestión de Clientes", use_container_width=True):
                    st.session_state.pantalla = "gestion_clientes"
                    st.rerun()
        with col_b4:
            if st.button("← Salir", use_container_width=True):
                st.session_state.pantalla = "login"
                st.rerun()

    st.markdown("<h3 style='margin:4px 0 0 0;'>Esquema comercial 2017</h3>", unsafe_allow_html=True)

    st.markdown("---")
    col_panel_filtros, col_panel_resultados = st.columns([3, 7])
    
    # PANEL LATERAL DE FILTROS Y SEGMENTACIÓN
    with col_panel_filtros:
        st.markdown("<div class='sombra-tenue'><h4 class='titulo-negro'>👤 Selección de Cliente</h4></div>", unsafe_allow_html=True)
        
        opciones_modo = ["Cliente Preexistente", "Cliente Nuevo", "Uso libre / Consulta"]
        val_actual = st.session_state.tipo_cliente_seleccion
        idx_m = opciones_modo.index(val_actual) if val_actual in opciones_modo else 0
        
        tipo_cli_sel = st.radio("Modo de atención:", options=opciones_modo, index=idx_m, key="radio_tipo_cliente")
        
        if tipo_cli_sel != st.session_state.tipo_cliente_seleccion:
            st.session_state.tipo_cliente_seleccion = tipo_cli_sel

        df_dir_clientes = cargar_catalogo_clientes()

        if tipo_cli_sel == "Cliente Preexistente":
            if df_dir_clientes.empty:
                st.caption("⚠️ No hay clientes en el directorio.")
            else:
                dict_clientes = {}
                opciones_combo = ["-- Seleccionar --"]
                for _, row in df_dir_clientes.iterrows():
                    nom, num = str(row['Nombre Cliente']).strip(), str(row['Numero Cliente']).strip()
                    etiqueta = f"{nom} {num}".strip()
                    opciones_combo.append(etiqueta)
                    dict_clientes[etiqueta] = (nom, num)
                
                cli_sel_str = st.selectbox("Selecciona el cliente:", options=opciones_combo)
                if cli_sel_str != "-- Seleccionar --":
                    nom_sel, num_sel = dict_clientes[cli_sel_str]
                    st.session_state.cliente_nombre = nom_sel
                    st.session_state.cliente_numero = num_sel
                    
                    ultima_aud = obtener_ultima_auditoria(nom_sel, num_sel)
                    if ultima_aud:
                        st.markdown(f"<div class='alerta-ultima-auditoria'>📌 <b>Última auditoría:</b> {ultima_aud}</div>", unsafe_allow_html=True)

        elif tipo_cli_sel == "Cliente Nuevo":
            st.session_state.cliente_nombre = st.text_input("Nombre del Cliente Nuevo:", value=st.session_state.cliente_nombre)
            st.session_state.cliente_numero = st.text_input("Número de Cliente:", value=st.session_state.cliente_numero)

        elif tipo_cli_sel == "Uso libre / Consulta":
            st.session_state.cliente_nombre = ""
            st.session_state.cliente_numero = ""

        if total_sel > 0:
            st.markdown("---")
            if st.button("💾 Confirmar y Guardar Visita", use_container_width=True, type="primary"):
                nombre_guardado = st.session_state.cliente_nombre or "Cliente General"
                if consolidar_y_guardar_visita_actual():
                    st.toast(f"✅ ¡Visita de '{nombre_guardado}' guardada exitosamente!")
                reiniciar_pantalla_total()
                st.rerun()

            if st.button("🧹 Limpiar Pantalla / Cancelar", use_container_width=True):
                reiniciar_pantalla_total()
                st.toast("🧹 Vista limpiada por completo.")
                st.rerun()

        st.markdown("---")
        st.markdown("<div class='sombra-tenue'><h4 class='titulo-negro'>Modo de Consulta</h4></div>", unsafe_allow_html=True)
        
        opciones_vista = ["🔎 Catálogo de Productos", "🏷️ Listado de Marcas", "🔍 Búsqueda"]
        idx_v = 0 if st.session_state.vista_catalogo == "productos" else (1 if st.session_state.vista_catalogo == "marcas" else (2 if st.session_state.vista_catalogo == "busqueda" else None))

        modo_sel = st.radio("Vista:", options=opciones_vista, index=idx_v, label_visibility="collapsed", key="radio_modo_consulta")
        
        nuevo_m = None
        if modo_sel:
            if "Productos" in modo_sel:
                nuevo_m = "productos"
            elif "Marcas" in modo_sel:
                nuevo_m = "marcas"
            else:
                nuevo_m = "busqueda"
            
        if nuevo_m != st.session_state.vista_catalogo:
            st.session_state.vista_catalogo = nuevo_m
            st.rerun()

        if st.session_state.vista_catalogo == "productos":
            st.markdown("<div class='sombra-tenue'><h4 class='titulo-negro'>Segmentación</h4></div>", unsafe_allow_html=True)
            
            if df_productos is not None:
                if columna_familia_real and columna_familia_real in df_productos.columns:
                    lista_familias = sorted(df_productos[columna_familia_real].dropna().unique().tolist())
                else:
                    posibles_cols = [c for c in df_productos.columns if "familia" in normalizar_texto(c)]
                    if posibles_cols:
                        columna_familia_real = posibles_cols[0]
                        lista_familias = sorted(df_productos[columna_familia_real].dropna().unique().tolist())
                    else:
                        lista_familias = []

                op_def = "-- Selecciona una familia --"
                fams = [op_def] + lista_familias
                
                idx_f = fams.index(st.session_state.filtro_familia) if st.session_state.filtro_familia in fams else 0
                sel_f = st.selectbox("Familia:", options=fams, index=idx_f, label_visibility="collapsed")
                
                if sel_f != st.session_state.filtro_familia:
                    st.session_state.filtro_familia = sel_f
                    st.rerun()

            if st.session_state.filtro_familia != "-- Selecciona una familia --" or st.session_state.busqueda_rapida != "":
                if st.button("🔄 Limpiar Filtros", use_container_width=True):
                    st.session_state.filtro_familia = "-- Selecciona una familia --"
                    st.session_state.busqueda_rapida = ""
                    st.rerun()

    # PANEL DE RESULTADOS Y PRODUCTOS
    with col_panel_resultados:
        if st.session_state.vista_catalogo is None:
            st.info("👈 **Por favor selecciona un Modo de Consulta** en el panel izquierdo.")
        
        elif st.session_state.vista_catalogo == "productos":
            if st.session_state.filtro_familia == "-- Selecciona una familia --":
                st.info("👈 **Por favor selecciona una familia** para consultar sus productos.")
            elif df_productos is None:
                st.error("⚠️ No se pudo cargar 'productos.csv'.")
            else:
                df_filtrado = df_productos.copy()
                
                busqueda = st.text_input("🔤 Búsqueda rápida:", value=st.session_state.busqueda_rapida, placeholder="Escribe código, clave o descripción...")

                if busqueda != st.session_state.busqueda_rapida:
                    st.session_state.busqueda_rapida = busqueda
                    st.rerun()

                fam_activa = st.session_state.filtro_familia != "-- Selecciona una familia --"
                if fam_activa and columna_familia_real and columna_familia_real in df_filtrado.columns:
                    df_filtrado = df_filtrado[df_filtrado[columna_familia_real] == st.session_state.filtro_familia]

                if st.session_state.busqueda_rapida.strip():
                    cond = pd.Series(False, index=df_filtrado.index)
                    for col_val in df_filtrado.columns:
                        cond |= df_filtrado[col_val].astype(str).str.contains(st.session_state.busqueda_rapida, case=False, na=False)
                    df_filtrado = df_filtrado[cond]

                st.markdown(
                    f"""<div style='text-align:center; margin: 10px 0 16px 0;'>
                        <span style='background-color:#F97316; color:white; font-weight:bold; padding:8px 22px; border-radius:20px; font-size:14px; display:inline-block;'>
                            🏷️ {st.session_state.filtro_familia if fam_activa else 'Todas'}
                        </span>
                    </div>""",
                    unsafe_allow_html=True
                )

                if not df_filtrado.empty:
                    cols_a_mostrar = [c for c in st.session_state.columnas_seleccionadas if c in df_filtrado.columns]
                    st.dataframe(df_filtrado[cols_a_mostrar], use_container_width=True, hide_index=True, height=260)
                    
                    st.markdown("---")
                    st.markdown("#### 📌 Marcar productos de esta vista:")
                    
                    for idx, row in df_filtrado.head(30).iterrows():
                        p_id = str(row.iloc[0]) if len(row) > 0 else f"PROD_{idx}"
                        
                        marcado = p_id in st.session_state.productos_seleccionados
                        ubi_act = st.session_state.productos_seleccionados[p_id]["ubicacion"] if marcado else "Piso de Venta"

                        cdet, cubi, cchk = st.columns([6, 3, 1])
                        with cdet:
                            desc = row.get("Descripcion de producto", str(row.iloc[1] if len(row)>1 else p_id))
                            st.write(f"**{p_id}** - {desc}")
                        with cubi:
                            u_sel = st.selectbox("Ubicación", ["Piso de Venta", "Almacén", "Ambos"], index=["Piso de Venta", "Almacén", "Ambos"].index(ubi_act), key=f"sel_ubic_{p_id}_{idx}", label_visibility="collapsed")
                        with cchk:
                            chk = st.checkbox("", value=marcado, key=f"chk_p_{p_id}_{idx}")

                        if chk:
                            st.session_state.productos_seleccionados[p_id] = {"datos": row.to_dict(), "ubicacion": u_sel}
                        elif p_id in st.session_state.productos_seleccionados:
                            del st.session_state.productos_seleccionados[p_id]

                        st.markdown("<hr style='margin:2px 0; border:0; border-top: 1px solid #E2E8F0;'>", unsafe_allow_html=True)
                else:
                    st.warning("⚠️ No se encontraron productos con los filtros seleccionados.")

        elif st.session_state.vista_catalogo == "marcas":
            st.markdown("<h4 style='color: #000000;'>🏷️ Listado de Marcas</h4>", unsafe_allow_html=True)
            st.markdown("---")
            if df_marcas is None:
                st.warning("⚠️ No se cargó 'marcas.csv' o 'marcas.xlsx'.")
            else:
                col_m = df_marcas.columns[0]
                for idx, row in df_marcas.iterrows():
                    nom_m = str(row[col_m]).strip()
                    if not nom_m or nom_m.lower() == "nan":
                        continue

                    marcado_m = nom_m in st.session_state.marcas_seleccionadas
                    ubi_m_act = st.session_state.marcas_seleccionadas[nom_m]["ubicacion"] if marcado_m else "Piso de Venta"

                    cnom, cubi, cchk = st.columns([6, 3, 1])
                    with cnom:
                        st.markdown(f"🏷️ **{nom_m}**")
                    with cubi:
                        u_m_sel = st.selectbox("Ubicación Marca", ["Piso de Venta", "Almacén", "Ambos"], index=["Piso de Venta", "Almacén", "Ambos"].index(ubi_m_act), key=f"sel_ubi_m_{nom_m}_{idx}", label_visibility="collapsed")
                    with cchk:
                        chk_m = st.checkbox("", value=marcado_m, key=f"chk_m_{nom_m}_{idx}")

                    if chk_m:
                        st.session_state.marcas_seleccionadas[nom_m] = {"ubicacion": u_m_sel}
                    elif nom_m in st.session_state.marcas_seleccionadas:
                        del st.session_state.marcas_seleccionadas[nom_m]

                    st.markdown("<hr style='margin:2px 0; border:0; border-top: 1px solid #E2E8F0;'>", unsafe_allow_html=True)

        elif st.session_state.vista_catalogo == "busqueda":
            st.markdown("<h4 style='color: #000000;'>🔍 Búsqueda</h4>", unsafe_allow_html=True)
            st.markdown("---")

            df_precios = cargar_lista_precios()

            if df_precios is None:
                st.warning("⚠️ No se cargó 'lista_precios.csv'.")
            else:
                columnas_busqueda = [c for c in ["código", "clave", "descripción", "Descripción Familia"] if c in df_precios.columns]
                columnas_visibles = [c for c in ["Descripción Familia", "código", "clave", "descripción"] if c in df_precios.columns]

                texto_busqueda = st.text_input("🔍 Búsqueda:", placeholder="Escribe descripción, código, clave o familia...")

                if texto_busqueda.strip():
                    mascara = pd.Series(False, index=df_precios.index)
                    for col in columnas_busqueda:
                        mascara = mascara | df_precios[col].astype(str).str.contains(texto_busqueda.strip(), case=False, na=False)
                    df_resultado_busqueda = df_precios[mascara]
                    if "Descripción Familia" in df_resultado_busqueda.columns:
                        df_resultado_busqueda = df_resultado_busqueda.sort_values(by="Descripción Familia", ascending=True)

                    st.caption(f"Coincidencias: `{len(df_resultado_busqueda)}`")
                    st.dataframe(df_resultado_busqueda[columnas_visibles], use_container_width=True, hide_index=True, height=400)
                else:
                    st.info("Escribe algo arriba para buscar en la lista de precios.")